import io
from celery import shared_task
from celery.contrib.abortable import AbortableTask
from django.core.cache import cache
import random
import redis
from celery import shared_task, current_task
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from django.conf import settings
import oracledb
import math
from datetime import datetime
import pandas as pd 

REQUIRED_COLUMNS = ["ARTICLE_CODE", "SU", "UNIQ_CODE", "UNIQ_NAME", "FROM_DATE", "TO_DATE","FLYER_RSP", "REG_RSP", "UNIT_DN", "REMARKS", "FLYER_TYPE", "CREATED_BY", "APPLICABLE_LOCATIONS"]
DATE_COLUMNS = ["FROM_DATE", "TO_DATE"]
REQUIRED_VALUE_COLUMNS = ["ARTICLE_CODE", "SU", "UNIQ_NAME", "UNIQ_CODE", "FROM_DATE", "TO_DATE", "FLYER_RSP", "FLYER_TYPE", "CREATED_BY"]
COLS_TO_FILL =[ "SU", "UNIQ_NAME", "FROM_DATE", "TO_DATE",  "REMARKS", "FLYER_TYPE", "CREATED_BY"]
INVALID_STRINGS = {"n/a", "na", "null", "none", "-"}
date_format = "%d-%b-%y"

r = redis.Redis(host='localhost', port=6379, db=0)





def clean_value(val):
    if isinstance(val, pd.Timestamp):  # convert pandas Timestamp to datetime
        return val.to_pydatetime()
    elif isinstance(val, datetime):  # datetime is fine as-is
        return val
    elif isinstance(val, float) and math.isnan(val):  # convert NaN to None
        return None
    else:
        return val


def connection():
        
        username = settings.ORACLE_USERNAME
        password = settings.ORACLE_PASSWORD
        dsn = settings.ORACLE_DSN
        client_path = settings.ORACLE_CLIENTPATH
        print(username)
        print(password)
        print(dsn)
        print(client_path)
        try:
      
            oracledb.init_oracle_client(lib_dir=client_path)
       
            conn = oracledb.connect(user=username, password=password, dsn=dsn)
       
            return conn
        except oracledb.Error as e:
           
            raise
        except Exception as e:
            raise



@shared_task(bind=True, queue='heavy_queue1', base=AbortableTask)
def verify_excel_file(self, cache_key, file_info, file_name, user_id):
    task_id = self.request.id
    clean_user_id = str(user_id).replace(" ", "_")
    user_group = f"notifications_{clean_user_id}"
    channel_layer = get_channel_layer()

    file_id = file_info.get("id")
    file_bytes = cache.get(cache_key)

    def send_message(status, messages, progress=0):
        # print("Sending message:", status, messages, progress)
        payload = {
            "type": "notification.update",
            "id": file_id,
            "task_id": task_id,
            "filename": file_name,
            "status": status,
            "progress": progress,
            "messages": messages
        }
        if channel_layer:
            async_to_sync(channel_layer.group_send)(user_group, payload)
        else:
            print("No channel layer or user offline, skipping message:", payload)
    # if not file_path or not os.path.exists(file_path):
    #     send_message(
    #         "error", "No file uploaded or file not found", progress=100)
    #     return
    send_message("apiUpdate", "Validation started", progress=0)
    try:
        # with open(file_path, "rb") as f:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    except (BadZipFile, InvalidFileException) as e:
        send_message(
            "error", f"Error loading Excel file: {str(e)}", progress=100)
        return

    except Exception as e:
        send_message(
            "error", f"Error reading Excel file with pandas: {str(e)}", progress=100)
        return

    else:
        send_message("success", "Excel file verified successfully", progress=1)

    try:
        all_sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
    except Exception as e:
        send_message(
            "error", f"Error reading Excel file with pandas: {str(e)}", 100)
        return

    print("here")
    notifications_count = 0
    all_sheets_list = list(all_sheets.items())
    total_sheets = len(all_sheets_list)

    section_dtls = []
    for s_idx, (sheet_name, df) in enumerate(all_sheets_list):
        sheet_base_prg = (s_idx / total_sheets) * 100
        sheet_end_prg = ((s_idx + 1) / total_sheets) * 100
        send_message("success", f"Starting {sheet_name} sheet", progress=int(sheet_base_prg))

        # Check if sheet is empty
        if df.empty:
            send_message("error", f"Sheet '{sheet_name}' is empty", int(sheet_base_prg + 2))
            notifications_count += 1
            continue

        # Check for missing required columns
        missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
        if missing_columns:
            send_message("error", f"Sheet '{sheet_name}': Missing required columns: {', '.join(missing_columns)}", int(sheet_base_prg + 4))
            notifications_count += 1
            continue  # Skip validation if columns missing

        # ---- Begin value validation ----
        rules = {
            'ARTICLE_CODE': 'number',
            'FLYER_RSP': 'number',
            'UNIT_DN': 'number',
            'CREATED_BY': 'string',
            'REG_RSP':'number'
        }

        messages = []
        for col, rule in rules.items():
            if col not in df.columns:
                continue  # skip columns missing from sheet
            
            if rule == 'number':
                mask = ~(pd.to_numeric(df[col], errors='coerce').notna() | df[col].isna())
            elif rule == 'string':
                mask = ~((df[col].apply(lambda x: isinstance(x, str)) | df[col].isna()))
            
            for idx in df[mask].index:
                messages.append(f"Row {idx + 2}, Column '{col}' has invalid value: {df.at[idx, col]}")

        if messages:
            msg_text = "\n".join(messages)
            send_message("error", f"Sheet '{sheet_name}' has invalid entries:\n{msg_text}", int(sheet_base_prg + 6))
            notifications_count += len(messages)
        else:
            send_message("success", f"Sheet '{sheet_name}' passed all value validations", int(sheet_base_prg + 6))

        def is_empty_row(row):
            return row.isna().all() or (row.astype(str).str.strip() == "").all()

        def is_last_item_only(row):
            """
            Returns True if all items in the row are null/empty 
            except for the very last column.
            """
            # Convert to a pandas Series if it isn't one already to use .isna()
            s = pd.Series(row)

            # prefix: all elements except the last one
            prefix = s.iloc[:-1]

            # last_item: the final element
            last_item = s.iloc[-1]

            # Condition 1: All elements in the prefix must be null
            prefix_is_empty = prefix.isna().all()

            # Condition 2: The last item must NOT be null
            last_is_not_empty = pd.notna(last_item)

            return prefix_is_empty and last_is_not_empty
        invalid_mask = df.copy()
        for col in df.columns:
            invalid_mask[col] = (
                df[col].isna() |
                (df[col].astype(str).str.strip() == "") |
                (df[col].astype(str).str.lower().isin(INVALID_STRINGS))
            )

        previous_row_empty = False
        total_rows = len(df)
        last_sent_progress = -1
        STEP = 19

        removable_rows = []
        is_last_item_only_row = []

        for i, row in df.iterrows():
            row_progress = int(
                sheet_base_prg + ((i / total_rows) *
                                  (sheet_end_prg - sheet_base_prg))
            )

            if row_progress >= last_sent_progress + STEP:
                last_sent_progress = row_progress
                send_message("success", "", progress=row_progress)
            if r.get(f"stop_{self.request.id}"):
                r.delete(f"stop_{self.request.id}")  # Cleanup
                send_message("error", "User stopped Validation manually", 100)
                return "User Stopped validation manually"
            excel_row = i + 2

            if is_empty_row(row):
                previous_row_empty = True

                print(f"row > {excel_row} empty row")
                removable_rows.append(excel_row)
                continue

            if is_last_item_only(row):
                print(f"row > {excel_row} last item only")

                if previous_row_empty:
                    notifications_count += 1
                    # send_message( "error",f"Sheet '{sheet_name}': "  f"Empty row found in {excel_row-1} ",progress=row_progress+ 3 )
                previous_row_empty = False
                is_last_item_only_row.append(excel_row-2)
                continue

            previous_row_empty = False

            for col in REQUIRED_VALUE_COLUMNS:

                if notifications_count > 500:
                    r.delete(f"stop_{self.request.id}")  # Cleanup
                    send_message(
                        "error", "process stoped Automatically because of 500 + errors  ", 100)
                    send_message("apiUpdate", "Validation Stopped", progress=100)
                    return "process stoped Automatically because of 500 + errors "

                if r.get(f"stop_{self.request.id}"):
                    r.delete(f"stop_{self.request.id}")  # Cleanup
                    send_message(
                        "error", "User stopped Validation manually", 100)
                    return "User Stopped validation manually"
                if col not in df.columns:
                    break

                ws = wb[sheet_name]
                excel_row = i + 2
                resolved_value = row[col]

                if invalid_mask.at[i, col]:

                    excel_col_idx = list(df.columns).index(col) + 1
                    cell_addr = f"{get_column_letter(excel_col_idx)}{excel_row}"

                    # ---- MERGE CHECK ----
                    is_merged = False

                    for merged_range in ws.merged_cells.ranges:
                        if cell_addr in merged_range:
                            is_merged = True
                            min_col, min_row, _, _ = merged_range.bounds
                            resolved_value = ws.cell(
                                row=min_row, column=min_col).value
                            break

                    if is_merged and resolved_value is not None:
                        # send_message("error", f"Sheet '{sheet_name}': Row {i+2}, Col '{col}' invalid", row_progress)
                        if col in DATE_COLUMNS:
                            try:
                                if pd.isna(resolved_value):
                                    raise ValueError("Empty date")

                                # ---- HANDLE TIMESTAMP / DATETIME ----
                                if isinstance(resolved_value, (pd.Timestamp, datetime)):
                                    date_obj = resolved_value.date()

                                else:
                                    # string or other → parse then take date
                                    date_obj = pd.to_datetime(
                                        resolved_value).date()

                                # ---- FINAL STRICT FORMAT CHECK ----
                                datetime.strptime(date_obj.strftime(
                                    date_format), date_format)

                                continue  # ✅ valid date

                            except Exception as e:
                                send_message(
                                    "error", f"Sheet '{sheet_name}': Invalid date in column {e}" f"'{col}', row {excel_row}: {resolved_value}", progress=row_progress+6)
                                notifications_count += 1
                                continue

                        continue  # merged + has value = valid
                    else:
                        send_message(
                            "error", f"Sheet '{sheet_name}': Row {excel_row}, "f"Col '{col}' is missing or invalid", progress=row_progress+10)
                        notifications_count += 1

                if r.get(f"stop_{self.request.id}"):
                    r.delete(f"stop_{self.request.id}")  # Cleanup
                    send_message(
                        "error", "User stopped Validation manually", 100)
                    return "User Stopped validation manually"

                if notifications_count > 500:
                    r.delete(f"stop_{self.request.id}")  # Cleanup
                    send_message(
                        "error", "process stoped Automatically because of 500 + errors  ", 100)
                    return "process stoped Automatically because of 500 + errors "
                
            print(f"row > {excel_row} validated")
        send_message(
            "success", f"Finished {sheet_name}", progress=int(sheet_end_prg))
        if  removable_rows:
            section_dtls.append({
                "sheetname": sheet_name,
                "removable_rows": removable_rows,
                "is_last_item_only_row" : is_last_item_only_row
            })

    cache.set(f"upload_key_{file_id}", section_dtls, timeout=3600)
    send_message("success", "File fully validated.", 100)
    
    send_message("warning", f"upload_key_{file_id}", 100)

    send_message("apiUpdate", "Validation completed", progress=100)

    print(section_dtls)

    return {"status": "Completed", "file": file_name}


@shared_task(bind=True, queue='heavy_queue1',    base=AbortableTask, autoretry_for=(Exception,), retry_backoff=5, retry_kwargs={'max_retries': 3})
def process_uploaded_file(self, cache_key, file_id=None, filename=None, task_id=None ,user_id=None , sectionKey=None):

    # 🔑 Load from cache
    file_bytes = cache.get(cache_key)
    all_sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
    all_sections = cache.get(sectionKey)

    all_sheets_list = list(all_sheets.items())
    progress_counter = 0
    START_ROW = 2
    all_rows_to_upload = []
    total_rows = sum(len(df) for _, df in all_sheets_list)
    processed_rows = 0
    cache_key=f"final-notification"
    max_records=50

    def print_progress(state, message, progress, counter):
        """
        Update progress details and store in cache with a maximum number of records.
        Old records are removed when the limit is reached.
        If progress is 100%, keep only the final record.
        """
        # Create the progress record
        progress_record = {
            "file_id": file_id,
            "filename": filename,
            "task_id": task_id,
            "state": state,
            "message": message,
            "progress": f"{progress:.0f}",
            "cache_key":cache_key
        }

        if progress >= 100:
            # If complete, keep only this final record
            cached_data = [progress_record]
        else:
            # Retrieve existing cache data (if any)
            cached_data = cache.get(cache_key, [])

            # Append the new progress record
            cached_data.append(progress_record)

            # Keep only the last `max_records` entries
            if len(cached_data) > max_records:
                # Separate records by task_id
                same_task = [r for r in cached_data if r["task_id"] == task_id]
                other_tasks = [r for r in cached_data if r["task_id"] != task_id]

                # Trim only the current task's records
                if len(same_task) > max_records:
                    same_task = same_task[-max_records:]

                # Recombine
                cached_data = other_tasks + same_task

        # Store updated data back to cache (expires in 1 hour)
        cache.set(cache_key, cached_data, timeout=60 * 10)
        print(max_records)
        # Increment counter
        counter += 1
        return counter
    for sheet_name, df in all_sheets_list:

        END_ROW = len(df) + 1  
         # Sheet started
        progress = (processed_rows / total_rows) * 100
        progress_counter = print_progress(
            state="INFO",
            message=f"Sheet '{sheet_name}' started",
            progress=progress,
            counter=progress_counter
        )
        breaks = next(
            (s["removable_rows"] for s in all_sections if s["sheetname"] == sheet_name),
            []
        )
        is_last_item_only_row = next(
            (item['is_last_item_only_row'] for item in all_sections if item['sheetname'] == sheet_name),
            []
        )

        sections = []
        current_start = START_ROW

        for b in breaks:
            section_end = b - 1
            if current_start <= section_end:
                sections.append((current_start, section_end))
            current_start = b + 1

        # final section (dynamic end)
        if current_start <= END_ROW:
            sections.append((current_start, END_ROW))

        for start, end in sections:
            progress = (processed_rows / total_rows) * 100
            progress_counter = print_progress(
                state="INFO",
                message=f"Sheet '{sheet_name}': Section start {start} → {end}",
                progress=progress,
                counter=progress_counter
            )
            cutted = df.iloc[start-2:end-1].copy()  # make a copy to safely modify

            locations = cutted["APPLICABLE_LOCATIONS"].dropna().astype(
                int).tolist()
            print("Locations:", locations)

            cutted = cutted.drop(columns=["APPLICABLE_LOCATIONS"])
            # keep rows NOT in the list
            mask = ~cutted.index.isin(is_last_item_only_row)
            filtered_cutted = cutted[mask]
            for col in COLS_TO_FILL:
                filtered_cutted[col] = filtered_cutted[col].ffill()
          
            for index, row in filtered_cutted.iterrows():
                for location in locations:
                    row_to_upload = row.copy()
                    row_to_upload["APPLICABLE_LOCATIONS"] = location
                    row_to_upload["UPLOADED_BY"] = user_id
                    row_to_upload["SHEET"] = sheet_name
                    all_rows_to_upload.append(row_to_upload)
                    
                # Section finished
                processed_rows += 1
                progress = (processed_rows / total_rows) * 100
                progress_counter = print_progress(
                    state="SUCCESS",
                    message=f"Sheet '{sheet_name}', row {index} processed",
                    progress=progress,
                    counter=progress_counter
                )

            # Section finished
            progress = (processed_rows / total_rows) * 100
            progress_counter = print_progress(
                state="INFO",
                message=f"Sheet '{sheet_name}': Section end {start} → {end}",
                progress=progress,
                counter=progress_counter
            )
        progress = (processed_rows / total_rows) * 100
        progress_counter = print_progress(
            state="INFO",
            message=f"Sheet '{sheet_name}' finished",
            progress=progress,
            counter=progress_counter
        )


       
    conn = connection()
    cursor = conn.cursor()

    if all_rows_to_upload:  # make sure this list is not empty
        columns = list(all_rows_to_upload[0].keys())
        placeholders = ", ".join([f":{i+1}" for i in range(len(columns))])
        
        sql = f"""
        INSERT INTO KWT_FLYER_ART_UPLOAD_WKLY
        ({', '.join(columns)})
        VALUES ({placeholders})
        """
        db_total = len(all_rows_to_upload)
        db_processed = 0


        try:
            for i, row in enumerate(all_rows_to_upload, start=1):
                values = [clean_value(row[col]) for col in columns]
                cursor.execute(sql, values)
                db_processed += 1

                progress = (db_processed / db_total) * 100
                progress_counter = print_progress(
                    state="SUCCESS",
                    message=f"Inserting row {db_processed} into DB . total {db_total} rows .  current sheet {row["SHEET"]}",
                    progress=progress,
                    counter=progress_counter
                )

            conn.commit()
            progress_counter = print_progress(
                state="SUCCESS",
                message="All rows inserted into DB",
                progress=100,
                counter=progress_counter
            )

        except Exception as e:
            conn.rollback()
            print("Error inserting rows:", e)

        finally:
            cursor.close()
            conn.close()

    
    
    progress_counter = print_progress(
        state="SUCCESS",
        message=f"Task completed for file '{filename}' (This message will be removed after one hour)",
        progress=100,
        counter=progress_counter
    )

    return {
        "rows": "f",
        "filename": filename,
        "task_id": task_id
    }
