import redis
import json
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
import pandas as pd
from datetime import datetime
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import uuid
from django.conf import settings
from django.http import JsonResponse
from apps.flareupload.tasks import verify_excel_file ,process_uploaded_file

REQUIRED_COLUMNS = ["ARTICLE_CODE", "SU", "UNIQ_CODE", "UNIQ_NAME", "FROM_DATE", "TO_DATE",
                    "FLYER_RSP", "REG_RSP", "UNIT_DN", "REMARKS", "FLYER_TYPE", "CREATED_BY", "APPLICABLE_LOCATIONS"]
DATE_COLUMNS = ["FROM_DATE", "TO_DATE"]
REQUIRED_VALUE_COLUMNS = ["ARTICLE_CODE", "SU", "UNIQ_NAME", "UNIQ_CODE",
                          "FROM_DATE", "TO_DATE", "FLYER_RSP", "FLYER_TYPE", "CREATED_BY"]
INVALID_STRINGS = {"n/a", "na", "null", "none", "-"}
messages = []
date_format = "%d-%b-%y"


from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
from django.http import JsonResponse
import os, uuid, json

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import io

@api_view(['POST'])
def verify_file(request):
    uploaded_file = request.FILES.get('file')
    raw_data = request.POST.get("data")

   
    # 1️⃣ File existence check
    if not uploaded_file:
        return Response(
            {"success": False, "messages": [{"type": "error", "message": "No file provided"}]},
            status=status.HTTP_400_BAD_REQUEST
        )

    # 2️⃣ Extension check
    file_ext = os.path.splitext(uploaded_file.name)[1].lower()
    if file_ext != ".xlsx":
        return Response(
            {
                "success": False,
                "messages": [{"type": "error", "message": "Only .xlsx Excel files are allowed"}]
            },
            status=status.HTTP_400_BAD_REQUEST
        )



    file_name = f"{uuid.uuid4()}{file_ext}"
 

    file_bytes = uploaded_file.read()
    cache_key = f"excel:{uuid.uuid4()}"

    cache.set(cache_key, file_bytes, timeout=60 * 60)

    file_bytes = cache.get(cache_key)


    try:
        load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    except InvalidFileException:
        return Response(
            {
                "success": False,
                "messages": [{"type": "error", "message": "Invalid XLSX file"}]
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        return Response(
            {
                "success": False,
                "messages": [{"type": "error", "message": f"{str(e)}"}]
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # 5️⃣ Parse JSON data
    parsed_data = {}
    if raw_data:
        try:
            parsed_data = json.loads(raw_data)
        except json.JSONDecodeError:
            return Response(
                {
                    "success": False,
                    "messages": [{"type": "error", "message": "Invalid JSON in 'data' field"}]
                },
                status=status.HTTP_400_BAD_REQUEST
            )

    file_info = parsed_data.get("fileToUpload", {})
    file_id = file_info.get("id")

    # 6️⃣ Only NOW submit Celery task
    task = verify_excel_file.delay(
        cache_key,
        file_info,
        uploaded_file.name,
        12345
    )

    return Response(
        {
            "success": True,
            "id": file_id,
            "filename": uploaded_file.name,
            "savedfilename": file_name,
            "task_id": task.id,
            "messages": [{"type": "success", "message": "Wait for validation"}],
        },
        status=status.HTTP_200_OK
    )




r = redis.Redis(host='localhost', port=6379, db=0)


def stop_task(request, task_id):
    if not task_id:
        return JsonResponse({'status': 'Task closed'}, status=200)
    r.set(f"stop_{task_id}", "true", ex=3600)
    return JsonResponse({'status': 'Stop signal sent to Redis'})


from django.core.cache import cache
import uuid

@api_view(['POST'])
def finalize_upload(request):
    files = request.FILES.getlist('files')
    task_ids = request.POST.getlist('task_ids')
    sectionKeys = request.POST.getlist('sectionKeys')
    print(task_ids)

    for uploaded_file, task_id ,sectionKey in zip(files, task_ids ,sectionKeys):
        print(sectionKey)
        # 🔹 uploaded_file is InMemoryUploadedFile
        file_bytes = uploaded_file.read()

        cache_key = f"excel:{uuid.uuid4()}"

        # Store in Redis (RAM)
        cache.set(cache_key, file_bytes, timeout=60 * 60)

        tesk = process_uploaded_file.delay(
            cache_key=cache_key,
            filename=uploaded_file.name,
            task_id=task_id,
            user_id=12345,
            sectionKey =sectionKey
        )

    return Response({"status": "queued"})