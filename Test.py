import pandas as pd
from datetime import datetime
# Define your required columns
REQUIRED_COLUMNS = [ "ARTICLE_CODE", "SU", "UNIQ_CODE", "UNIQ_NAME", "FROM_DATE", "TO_DATE", "FLYER_RSP", "REG_RSP", "UNIT_DN", "REMARKS", "FLYER_TYPE", "CREATED_BY", "APPLICABLE_LOCATIONS"]
DATE_COLUMNS = ["FROM_DATE", "TO_DATE"]
REQUIRED_VALUE_COLUMNS = ["ARTICLE_CODE", "SU", "UNIQ_NAME","UNIQ_CODE", "FROM_DATE", "TO_DATE", "FLYER_RSP", "FLYER_TYPE", "CREATED_BY"]
INVALID_STRINGS = {"n/a", "na", "null", "none", "-"}
date_format = "%d-%b-%y"
from openpyxl import load_workbook
path = r"C:\Users\HP\Downloads\testinh\1 page ok.xlsx"
# path = "TestFlareExcel.xlsx"
all_sheets = pd.read_excel(path, sheet_name=None) 
all_sheets = {"Sheet1": all_sheets["Sheet1"]}  # For testing, only use "Sheet1"
wb = load_workbook(path, data_only=True)
file_info = {"id": 222}
messages = []
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def send_message(status, messages, progress=0):
    # print(f"{status.upper()}: {messages} (Progress: {progress}%)")
    return
print("here")
notifications_count = 0
all_sheets_list = list(all_sheets.items())
total_sheets = len(all_sheets_list)
for s_idx, (sheet_name, df) in enumerate(all_sheets_list):
    # all_valid = (pd.to_numeric(df['UNIT_DN'], errors='coerce').notna() | df['UNIT_DN'].isna()).all()
    # print(all_valid)


    rules = {
    'ARTICLE_CODE': 'number',
    'FLYER_RSP': 'number',
    'UNIT_DN': 'number',
    'CREATED_BY': 'string',
    'REG_RSP' : 'number'
    }

    # List to collect messages
    messages = []

    for col, rule in rules.items():
        if rule == 'number':
            mask = ~(pd.to_numeric(df[col], errors='coerce').notna() | df[col].isna())
        elif rule == 'string':
            mask = ~((df[col].apply(lambda x: isinstance(x, str)) | df[col].isna()))
        
        # Create readable messages
        for idx in df[mask].index:
            messages.append(f"Row {idx + 2}, Column '{col}' has invalid value: {df.at[idx, col]}")

    # Print all messages
    if messages:
        print("Invalid entries found:\n")
        for msg in messages:
            print(msg)
    else:
        print("All entries are valid.")

    sheet_base_prg = (s_idx / total_sheets) * 100
    sheet_end_prg = ((s_idx + 1) / total_sheets) * 100
    send_message("success", f"Starting {sheet_name} sheet" , progress=int(sheet_base_prg))
    if df.empty:
        send_message("error", f"Sheet '{sheet_name}' is empty", int(sheet_base_prg + 2))
        notifications_count +=1 
        hasErrors = True
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        send_message( "error", f"Sheet '{sheet_name}': Missing required columns: {', '.join(missing_columns)}",int(sheet_base_prg + 4)  )
        notifications_count +=1 
        hasErrors = True

    def is_empty_row(row):
        return row.isna().all() or (row.astype(str).str.strip() == "").all()
    
    def is_last_item_only(row):
        """
        Returns True if all items in the row are null/empty 
        except for the very last column.
        """
        # Convert to a pandas Series if it isn't one already to use .isna()
        s = pd.Series(row)
        
        # prefix: all elements except the last one
        prefix = s.iloc[:-1]
        
        # last_item: the final element
        last_item = s.iloc[-1]
        
        # Condition 1: All elements in the prefix must be null
        prefix_is_empty = prefix.isna().all()
        
        # Condition 2: The last item must NOT be null
        last_is_not_empty = pd.notna(last_item)
        
        return prefix_is_empty and last_is_not_empty
    invalid_mask = df.copy()
    for col in df.columns:
        invalid_mask[col] = (
            df[col].isna() | 
            (df[col].astype(str).str.strip() == "") | 
            (df[col].astype(str).str.lower().isin(INVALID_STRINGS))
        )

    previous_row_empty = False
    total_rows = len(df)
    last_sent_progress = -1
    STEP = 19

    for i, row in df.iterrows():
        row_progress = int(
            sheet_base_prg + ((i / total_rows) * (sheet_end_prg - sheet_base_prg))
        )

        if row_progress >= last_sent_progress + STEP:
            last_sent_progress = row_progress
            send_message("success", "", progress=row_progress)
     
        excel_row = i + 2 


        if is_empty_row(row):
            previous_row_empty = True
            print(f"row > {excel_row} empty row")
            continue

        if is_last_item_only(row):
            print(f"row > {excel_row} last item only")

            if previous_row_empty:
                notifications_count +=1 
            previous_row_empty = False
            print(f"row > {excel_row} last item only - skipped")
            continue

        previous_row_empty = False
        
        for col in REQUIRED_VALUE_COLUMNS:
            
            if col not in df.columns:
                break

            ws = wb[sheet_name]
            excel_row = i + 2
            resolved_value = row[col]

            if invalid_mask.at[i, col]:

                
                excel_col_idx = list(df.columns).index(col) + 1
                cell_addr = f"{get_column_letter(excel_col_idx)}{excel_row}"

                # ---- MERGE CHECK ----
                is_merged = False
                

                for merged_range in ws.merged_cells.ranges:
                    if cell_addr in merged_range:
                        is_merged = True
                        min_col, min_row, _, _ = merged_range.bounds
                        resolved_value = ws.cell(row=min_row, column=min_col).value
                        break

                

        
                if is_merged and resolved_value is not None:
                    if col in DATE_COLUMNS:
                        try:
                            if pd.isna(resolved_value):
                                raise ValueError("Empty date")

                            # ---- HANDLE TIMESTAMP / DATETIME ----
                            if isinstance(resolved_value, (pd.Timestamp, datetime)):
                                date_obj = resolved_value.date()

                            else:
                                # string or other → parse then take date
                                date_obj = pd.to_datetime(resolved_value).date()

                            # ---- FINAL STRICT FORMAT CHECK ----
                            datetime.strptime(date_obj.strftime(date_format), date_format)

                            continue  # ✅ valid date

                        except Exception as e: 
                            send_message( "error", f"Sheet '{sheet_name}': Invalid date in column {e}" f"'{col}', row {excel_row}: {resolved_value}", progress=row_progress+6)        
                            notifications_count +=1    
                            hasErrors =True
                            continue

                    continue  # merged + has value = valid
                else:
                    send_message( "error", f"Sheet '{sheet_name}': Row {excel_row}, "f"Col '{col}' is missing or invalid",progress=row_progress+10)
                    hasErrors = True
                    notifications_count +=1     

            
send_message("success", "File fully validated", 98)
send_message("apiUpdate", "Validation completed .", progress=98)

